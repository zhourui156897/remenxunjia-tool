import os
import sys
import re
from datetime import datetime
from typing import Optional

from flask import Flask, render_template, request, jsonify
from PIL import Image, ImageEnhance, ImageOps
import openpyxl


def _base_path():
    """Return the base path for bundled resources (templates, static)."""
    if getattr(sys, 'frozen', False):
        return sys._MEIPASS
    return os.path.dirname(os.path.abspath(__file__))


def _data_path():
    """Return a writable directory next to the executable (or project root)."""
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


app = Flask(
    __name__,
    template_folder=os.path.join(_base_path(), 'templates'),
    static_folder=os.path.join(_base_path(), 'static'),
)
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024
UPLOAD_FOLDER = os.path.join(_data_path(), 'uploads')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# ---------------------------------------------------------------------------
# OCR: extract stock names from the "热门询价" screenshot
# ---------------------------------------------------------------------------

try:
    import pytesseract

    if getattr(sys, 'frozen', False):
        _tess_dir = os.path.join(os.path.dirname(sys.executable), 'tesseract')
        _tess_exe = os.path.join(_tess_dir, 'tesseract.exe')
        if os.path.isfile(_tess_exe):
            pytesseract.pytesseract.tesseract_cmd = _tess_exe
            os.environ['TESSDATA_PREFIX'] = _tess_dir
    HAS_TESSERACT = True
except ImportError:
    HAS_TESSERACT = False

TESSERACT_ERROR = ''

from PIL import ImageChops

NOISE_KEYWORDS = {
    '热门询价', '放量跟踪', '序号', '标的', '现价', '涨跌幅',
    '前一日总市值', '前一日成交额', '操作', '去询价', '去跟踪',
    '亿', 'SH', 'SZ', 'CNTGROUP',
}


def _preprocess_variants(img: Image.Image) -> list[Image.Image]:
    """Generate multiple preprocessed images for multi-pass OCR."""
    w, h = img.size
    left_half = img.crop((0, 0, w // 2, h))
    scale = 3
    scaled = left_half.resize((left_half.width * scale, left_half.height * scale), Image.LANCZOS)

    rgb = scaled.convert('RGB')
    r, g, b = rgb.split()
    bright = ImageChops.lighter(r, ImageChops.lighter(g, b))

    variants = []

    # V1: max-channel, threshold 90 (good for colored text like 中兵红箭)
    inv1 = ImageOps.invert(bright)
    enh1 = ImageEnhance.Contrast(inv1).enhance(2.0)
    variants.append(enh1.point(lambda x: 255 if x > 90 else 0))

    # V2: max-channel, lower threshold 70 (catches faint text like 海格通信)
    variants.append(enh1.point(lambda x: 255 if x > 70 else 0))

    # V3: grayscale enhanced (good for white text like 中际旭创, 深信服)
    gray = scaled.convert('L')
    inv3 = ImageOps.invert(gray)
    enh3 = ImageEnhance.Contrast(inv3).enhance(2.5)
    variants.append(enh3.point(lambda x: 255 if x > 100 else 0))

    # V4: red channel (good for red-highlighted text)
    inv4 = ImageOps.invert(r)
    enh4 = ImageEnhance.Contrast(inv4).enhance(3.0)
    variants.append(enh4.point(lambda x: 255 if x > 110 else 0))

    return variants


def _parse_numbered_rows(text: str) -> dict[int, list[str]]:
    """Parse OCR text, using row numbers (01, 02...) as anchors.

    Returns {row_number: [candidate_chinese_names]}.
    """
    rows: dict[int, list[str]] = {}
    for line in text.splitlines():
        line = line.strip()
        if not line:
            continue
        m = re.match(r'^[^0-9]*(\d{1,2})\s+(.+?)(?:\s+\d)', line)
        if not m:
            continue
        row_num = int(m.group(1))
        if row_num < 1 or row_num > 50:
            continue
        raw = m.group(2).strip()
        raw = re.sub(r'["\'"=|*«»\[\]()（）]', '', raw)
        chinese_parts = re.findall(r'[\u4e00-\u9fff]{2,}', raw)
        for part in chinese_parts:
            if len(part) >= 2 and not any(kw in part for kw in NOISE_KEYWORDS):
                rows.setdefault(row_num, []).append(part)
    return rows


def ocr_extract_stock_names(image_path: str) -> list[str]:
    """Multi-pass OCR: run 4 preprocessing variants, fuse by row number."""
    global TESSERACT_ERROR
    if not HAS_TESSERACT:
        TESSERACT_ERROR = 'pytesseract 未安装'
        return []

    try:
        img = Image.open(image_path)
        variants = _preprocess_variants(img)
    except Exception as e:
        TESSERACT_ERROR = f'图片预处理失败: {e}'
        return []

    all_rows: dict[int, list[str]] = {}

    tessdata_dir = None
    if getattr(sys, 'frozen', False):
        candidate = os.path.join(os.path.dirname(sys.executable), 'tesseract', 'tessdata')
        if os.path.isdir(candidate):
            tessdata_dir = candidate

    try:
        for variant in variants:
            ocr_config = '--psm 6'
            if tessdata_dir:
                ocr_config += f' --tessdata-dir "{tessdata_dir}"'
            text = pytesseract.image_to_string(
                variant, lang='chi_sim+eng', config=ocr_config
            )
            parsed = _parse_numbered_rows(text)
            for num, names in parsed.items():
                all_rows.setdefault(num, []).extend(names)
    except Exception as e:
        TESSERACT_ERROR = f'Tesseract OCR 执行失败: {e}'
        return []

    TESSERACT_ERROR = ''
    final_names = []
    seen = set()
    for num in sorted(all_rows.keys()):
        candidates = all_rows[num]
        if not candidates:
            continue
        best = _pick_best_candidate(candidates)
        if best and best not in seen:
            seen.add(best)
            final_names.append(best)

    return final_names


def _pick_best_candidate(candidates: list[str]) -> Optional[str]:
    """Pick the best stock name from multiple OCR candidates for one row."""
    if not candidates:
        return None

    scored: dict[str, float] = {}
    for c in candidates:
        freq = candidates.count(c)
        clen = len(c)
        all_chinese = all('\u4e00' <= ch <= '\u9fff' for ch in c)
        score = freq * 10 + clen + (5 if all_chinese else 0)
        if c not in scored or score > scored[c]:
            scored[c] = score

    return max(scored, key=scored.get)


# ---------------------------------------------------------------------------
# Excel: parse option pricing from 香草看涨报价 sheet
# ---------------------------------------------------------------------------

SHEET_CANDIDATES = ['香草看涨', '香草看涨报价']


def parse_excel(file_path: str) -> dict:
    """
    Returns {
        'by_name': { '证券简称': {code, name, call_1m, call_2m} },
        'by_code': { '证券代码': {code, name, call_1m, call_2m} },
        'date': '2026-03-03'
    }
    Auto-detects sheet name, header row, and column layout.
    """
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

    result_by_name = {}
    result_by_code = {}
    date_str = ''

    ws = None
    for candidate in SHEET_CANDIDATES:
        if candidate in wb.sheetnames:
            ws = wb[candidate]
            break

    if ws is None:
        wb.close()
        return {'by_name': {}, 'by_code': {}, 'date': ''}

    def fmt(v):
        if v is None or v == '-' or v == '':
            return None
        try:
            val = float(v)
            if val < 1:
                return round(val * 100, 2)
            return round(val, 2)
        except (ValueError, TypeError):
            return None

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return {'by_name': {}, 'by_code': {}, 'date': ''}

    for cell in rows[0]:
        if hasattr(cell, 'strftime'):
            date_str = cell.strftime('%Y-%m-%d')
            break

    header_row_idx = None
    for i, row in enumerate(rows):
        for cell in (row or []):
            if cell and str(cell).strip() == '证券代码':
                header_row_idx = i
                break
        if header_row_idx is not None:
            break

    if header_row_idx is None:
        header_row_idx = 1

    headers = rows[header_row_idx] if header_row_idx < len(rows) else ()

    code_col = 0
    name_col = 1
    col_1m = None
    col_2m = None

    for ci, h in enumerate(headers):
        if h is None:
            continue
        s = str(h).strip().upper().replace('\uff08', '(').replace('\uff09', ')')
        if s == '证券代码':
            code_col = ci
        elif s == '证券简称':
            name_col = ci
        elif col_1m is None and '1M' in s and '100' in s:
            col_1m = ci
        elif col_2m is None and '2M' in s and '100' in s:
            col_2m = ci

    if col_1m is None:
        col_1m = name_col + 1
    if col_2m is None:
        col_2m = col_1m + 1

    data_start = header_row_idx + 1
    for row in rows[data_start:]:
        if not row or len(row) <= max(code_col, name_col):
            continue
        code = row[code_col]
        name = row[name_col]
        if not code or not name:
            continue
        call_1m = row[col_1m] if len(row) > col_1m else None
        call_2m = row[col_2m] if len(row) > col_2m else None

        entry = {
            'code': str(code),
            'name': str(name),
            'call_1m': fmt(call_1m),
            'call_2m': fmt(call_2m),
        }
        result_by_name[str(name)] = entry
        result_by_code[str(code)] = entry

    if not date_str:
        basename = os.path.basename(file_path)
        m = re.search(r'(\d{4}-\d{1,2}-\d{1,2})', basename)
        if m:
            date_str = m.group(1)
        else:
            date_str = datetime.now().strftime('%Y-%m-%d')

    return {'by_name': result_by_name, 'by_code': result_by_code, 'date': date_str}


def _edit_distance(a: str, b: str) -> int:
    """Compute Levenshtein distance between two strings."""
    m, n = len(a), len(b)
    dp = list(range(n + 1))
    for i in range(1, m + 1):
        prev = dp[0]
        dp[0] = i
        for j in range(1, n + 1):
            tmp = dp[j]
            if a[i - 1] == b[j - 1]:
                dp[j] = prev
            else:
                dp[j] = 1 + min(prev, dp[j], dp[j - 1])
            prev = tmp
    return dp[n]


def match_stocks(ocr_names: list[str], excel_data: dict) -> list[dict]:
    """Match OCR names against Excel data, return list of stock entries."""
    by_name = excel_data['by_name']
    results = []

    for ocr_name in ocr_names:
        matched = None
        if ocr_name in by_name:
            matched = by_name[ocr_name]
        else:
            for excel_name, entry in by_name.items():
                if ocr_name in excel_name or excel_name in ocr_name:
                    matched = entry
                    break

        if not matched:
            best_dist = 999
            best_entry = None
            for excel_name, entry in by_name.items():
                if abs(len(ocr_name) - len(excel_name)) > 2:
                    continue
                dist = _edit_distance(ocr_name, excel_name)
                threshold = 1 if len(ocr_name) <= 3 else 2
                if dist <= threshold and dist < best_dist:
                    best_dist = dist
                    best_entry = entry
            if best_entry:
                matched = best_entry

        if matched:
            results.append({
                'name': matched['name'],
                'code': matched['code'],
                'call_1m': matched['call_1m'],
                'call_2m': matched['call_2m'],
                'matched': True,
            })
        else:
            results.append({
                'name': ocr_name,
                'code': '',
                'call_1m': None,
                'call_2m': None,
                'matched': False,
            })

    return results


# ---------------------------------------------------------------------------
# Flask Routes
# ---------------------------------------------------------------------------

@app.route('/')
def index():
    today = datetime.now().strftime('%Y-%m-%d')
    return render_template('index.html', today=today)


@app.route('/upload', methods=['POST'])
def upload():
    screenshot = request.files.get('screenshot')
    excel_file = request.files.get('excel')
    manual_names_raw = request.form.get('manual_names', '')

    if not excel_file:
        return jsonify({'error': '请上传 Excel 报价表'}), 400

    excel_path = os.path.join(UPLOAD_FOLDER, 'pricing.xlsx')
    excel_file.save(excel_path)
    excel_data = parse_excel(excel_path)

    ocr_names = []
    if screenshot:
        img_path = os.path.join(UPLOAD_FOLDER, 'screenshot.png')
        screenshot.save(img_path)
        ocr_names = ocr_extract_stock_names(img_path)

    manual_names = []
    if manual_names_raw.strip():
        for part in re.split(r'[,，\n\r;；、\s]+', manual_names_raw):
            part = part.strip()
            if part and len(part) >= 2:
                manual_names.append(part)

    combined_names = ocr_names + [n for n in manual_names if n not in ocr_names]

    if combined_names:
        results = match_stocks(combined_names, excel_data)
    else:
        results = []

    all_stocks = [
        {'code': e['code'], 'name': e['name']}
        for e in excel_data['by_name'].values()
    ]

    return jsonify({
        'stocks': results,
        'date': excel_data['date'],
        'ocr_names': ocr_names,
        'manual_names': manual_names,
        'all_stocks': all_stocks,
        'has_tesseract': HAS_TESSERACT,
        'tesseract_error': TESSERACT_ERROR,
    })


@app.route('/lookup', methods=['POST'])
def lookup():
    """Look up a single stock by name in a previously uploaded Excel."""
    data = request.get_json()
    stock_name = data.get('name', '')
    excel_path = os.path.join(UPLOAD_FOLDER, 'pricing.xlsx')

    if not os.path.exists(excel_path):
        return jsonify({'error': 'Excel 文件未上传'}), 400

    excel_data = parse_excel(excel_path)
    by_name = excel_data['by_name']

    matched = None
    if stock_name in by_name:
        matched = by_name[stock_name]
    else:
        for ename, entry in by_name.items():
            if stock_name in ename or ename in stock_name:
                matched = entry
                break

    if not matched:
        best_dist = 999
        for ename, entry in by_name.items():
            if abs(len(stock_name) - len(ename)) > 2:
                continue
            dist = _edit_distance(stock_name, ename)
            threshold = 1 if len(stock_name) <= 3 else 2
            if dist <= threshold and dist < best_dist:
                best_dist = dist
                matched = entry

    if matched:
        return jsonify({
            'code': matched['code'],
            'name': matched['name'],
            'call_1m': matched['call_1m'],
            'call_2m': matched['call_2m'],
            'matched': True,
        })
    return jsonify({'matched': False})


@app.route('/batch_lookup', methods=['POST'])
def batch_lookup():
    """Look up multiple stocks by name, returning matched results."""
    data = request.get_json()
    names_raw = data.get('names', '')
    excel_path = os.path.join(UPLOAD_FOLDER, 'pricing.xlsx')

    if not os.path.exists(excel_path):
        return jsonify({'error': 'Excel 文件未上传'}), 400

    names = []
    for part in re.split(r'[,，\n\r;；、\s]+', names_raw):
        part = part.strip()
        if part and len(part) >= 2:
            names.append(part)

    if not names:
        return jsonify({'error': '请输入至少一个标的名称'}), 400

    excel_data = parse_excel(excel_path)
    results = match_stocks(names, excel_data)

    return jsonify({
        'stocks': results,
        'date': excel_data['date'],
    })


if __name__ == '__main__':
    import webbrowser
    import threading

    is_frozen = getattr(sys, 'frozen', False)

    def _open_browser():
        import time
        time.sleep(1.5)
        webbrowser.open('http://127.0.0.1:5050')

    if is_frozen:
        threading.Thread(target=_open_browser, daemon=True).start()

    app.run(debug=not is_frozen, host='127.0.0.1', port=5050)
